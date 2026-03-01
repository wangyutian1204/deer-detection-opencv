import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def export_excel(data, save_path):
    """导出Excel报告（无外部依赖）"""
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "鹿群计数结果"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ["处理时间", "原始文件路径", "计数结果（只）"]
    if len(data[0]) == 4:
        headers.append("处理后文件路径")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # 写入数据
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = align_center

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 50

    # 保存文件
    wb.save(save_path)
    print(f"Excel报告已保存至：{save_path}")